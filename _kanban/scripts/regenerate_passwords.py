"""Regenerate passwords for all real users.

Selects users from `users` table, applies exclusion rules, generates
new 10-char lowercase+digit passwords, hashes with sha256, updates the
DB in a single transaction, and exports results to JSON + XLSX.

Run:
    cd /srv/project/backend_main_node
    source venv/bin/activate
    python _kanban/scripts/regenerate_passwords.py
"""
from __future__ import annotations

import hashlib
import json
import secrets
import string
import sys
from pathlib import Path

# Allow running as a script — make the project root importable.
PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from sqlalchemy.orm import Session  # noqa: E402

from database.database import SessionLocal  # noqa: E402
from models.user import User  # noqa: E402
from models.employees import Employees  # noqa: E402

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
EXPORT_DIR = PROJECT_ROOT / "_kanban"
JSON_PATH = EXPORT_DIR / "passwords_export_2026_05_26.json"
XLSX_PATH = EXPORT_DIR / "passwords_export_2026_05_26.xlsx"

PASSWORD_LENGTH = 10
PASSWORD_ALPHABET = string.ascii_lowercase + string.digits

# Hard-coded exclusion logins (case-insensitive)
EXCLUDE_LOGINS_EXACT = {"ofik", "admin", "administrator"}

# Substrings that mark a login as a system/test account
LOGIN_BAD_SUBSTRINGS = ("integrator", "интегратор", "iiko", "kwaaka")

# Substrings that mark a linked employee name as a system/test account
NAME_BAD_SUBSTRINGS = ("интегратор", "integrator", "kwaaka", "iiko")

# Generic ADM employee-name markers — these are skipped even though some real
# people (e.g. Akzhan, Amir) also have main_role_code = 'ADM' and must be kept.
ADM_GENERIC_NAME_PREFIXES = ("пользователь",)
ADM_GENERIC_NAME_EXACT = {"администратор"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def generate_password() -> str:
    """Return a cryptographically-random 10-char password (a-z 0-9)."""
    return "".join(secrets.choice(PASSWORD_ALPHABET) for _ in range(PASSWORD_LENGTH))


def hash_password(plain: str) -> str:
    """Project-standard sha256 hex hash."""
    return hashlib.sha256(plain.encode()).hexdigest()


def employee_name_is_generic_adm(name: str | None) -> bool:
    if not name:
        # ADM with no name → treat as generic system account.
        return True
    n = name.strip().lower()
    if n in ADM_GENERIC_NAME_EXACT:
        return True
    if any(n.startswith(p) for p in ADM_GENERIC_NAME_PREFIXES):
        return True
    return False


def classify_user(user: User, emp: Employees | None) -> tuple[bool, str]:
    """Return (should_skip, reason). Empty reason means user is included."""
    login = (user.login or "").strip()
    login_l = login.lower()

    if not login:
        return True, "login is NULL/empty"

    if login_l in EXCLUDE_LOGINS_EXACT:
        return True, f"login '{login}' in hard-exclude list"

    for sub in LOGIN_BAD_SUBSTRINGS:
        if sub in login_l:
            return True, f"login contains '{sub}'"

    emp_name = (emp.name if emp else "") or ""
    emp_name_l = emp_name.lower()
    for sub in NAME_BAD_SUBSTRINGS:
        if sub in emp_name_l:
            return True, f"employee name contains '{sub}'"

    main_role_code = (emp.main_role_code if emp else None) or ""
    if main_role_code.upper() == "ADM" and employee_name_is_generic_adm(emp_name):
        return True, f"generic ADM account (name='{emp_name}')"

    return False, ""


def write_json(rows: list[dict], path: Path) -> None:
    path.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "passwords"

    headers = ["user_id", "login", "password", "app_role", "fio", "main_role_code"]
    ws.append(headers)

    header_font = Font(bold=True)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    # Column widths
    widths = {"user_id": 10, "login": 38, "password": 14, "app_role": 14,
              "fio": 42, "main_role_code": 18}
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(h, 18)

    ws.freeze_panes = "A2"
    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> int:
    db: Session = SessionLocal()
    try:
        users: list[User] = db.query(User).order_by(User.id).all()

        # Build iiko_id → Employees map (one query).
        iiko_ids = [u.iiko_id for u in users if u.iiko_id]
        emp_map: dict[str, Employees] = {}
        if iiko_ids:
            for emp in db.query(Employees).filter(Employees.iiko_id.in_(iiko_ids)).all():
                emp_map[emp.iiko_id] = emp

        included: list[tuple[User, Employees | None]] = []
        excluded: list[tuple[User, str]] = []

        for u in users:
            emp = emp_map.get(u.iiko_id) if u.iiko_id else None
            skip, reason = classify_user(u, emp)
            if skip:
                excluded.append((u, reason))
            else:
                included.append((u, emp))

        # ------------------------------------------------------------------
        # Print exclusion list BEFORE updating.
        # ------------------------------------------------------------------
        print("=" * 80)
        print(f"TOTAL USERS IN DB: {len(users)}")
        print(f"EXCLUDED: {len(excluded)}")
        print(f"WILL UPDATE: {len(included)}")
        print("=" * 80)
        print("\nFULL EXCLUSION LIST:")
        for u, reason in excluded:
            print(f"  id={u.id:<5} login={u.login!r:<55} -- {reason}")
        print()

        # ------------------------------------------------------------------
        # Generate passwords and stage updates.
        # ------------------------------------------------------------------
        export_rows: list[dict] = []
        verify_samples: list[dict] = []  # first 5

        for idx, (u, emp) in enumerate(included):
            new_password = generate_password()
            new_hash = hash_password(new_password)
            old_hash = u.password

            if idx < 5:
                verify_samples.append({
                    "user_id": u.id,
                    "login": u.login,
                    "old_hash": old_hash,
                    "new_password_plain": new_password,
                    "new_hash": new_hash,
                    "same?": old_hash == new_hash,
                })

            u.password = new_hash

            export_rows.append({
                "user_id": u.id,
                "login": u.login,
                "password": new_password,
                "app_role": u.app_role or "",
                "fio": (emp.name if emp else "") or "",
                "main_role_code": (emp.main_role_code if emp else "") or "",
            })

        # ------------------------------------------------------------------
        # Print verification sample.
        # ------------------------------------------------------------------
        print("VERIFICATION SAMPLE (first 5 users — hashes should differ):")
        for s in verify_samples:
            print(f"  id={s['user_id']} login={s['login']}")
            print(f"     old_hash: {s['old_hash']}")
            print(f"     new_hash: {s['new_hash']}  (plain: {s['new_password_plain']})")
            print(f"     same? {s['same?']}")
        print()

        # ------------------------------------------------------------------
        # Commit.
        # ------------------------------------------------------------------
        try:
            db.commit()
            print(f"DB COMMIT OK — {len(included)} users updated.")
        except Exception as exc:
            db.rollback()
            print(f"DB COMMIT FAILED, rolled back: {exc}")
            raise

        # ------------------------------------------------------------------
        # Sort and export.
        # ------------------------------------------------------------------
        def sort_key(r: dict) -> tuple[str, str]:
            return (r["app_role"] or "", r["fio"] or "")

        export_rows.sort(key=sort_key)

        write_json(export_rows, JSON_PATH)
        write_xlsx(export_rows, XLSX_PATH)

        # ------------------------------------------------------------------
        # Final report.
        # ------------------------------------------------------------------
        print()
        print("=" * 80)
        print("DONE")
        print("=" * 80)
        print(f"Total users in DB        : {len(users)}")
        print(f"Users updated            : {len(included)}")
        print(f"Users excluded           : {len(excluded)}")
        print(f"Export JSON              : {JSON_PATH}")
        print(f"Export XLSX              : {XLSX_PATH}")
        print()
        print("Excluded breakdown by reason:")
        from collections import Counter
        for reason, count in Counter(r for _, r in excluded).most_common():
            print(f"  {count:>3}  {reason}")

        return 0
    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(main())
